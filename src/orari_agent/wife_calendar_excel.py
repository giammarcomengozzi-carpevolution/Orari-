"""Import Excel del calendario moglie.

L'unica informazione operativa estratta è la presenza del codice ``M``: tutte
le altre celle, codici e formattazioni vengono ignorati.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class WifeCalendarExcelImportResult:
    """Risultato dell'estrazione di date con codice M da un file XLSX."""

    dates: list[str]
    scanned_cells: int
    m_cells: int
    warnings: list[str]


def extract_m_dates_from_excel(path: str | Path) -> WifeCalendarExcelImportResult:
    """Estrae solo le date collegate a celle con codice esatto ``M``.

    Il parser è volutamente prudente: scansiona tutti i fogli, ignora colori e
    formattazione, accetta date reali Excel o stringhe ISO/date italiane e, per
    ogni cella ``M``, cerca una data nella stessa riga o nella stessa colonna.
    """

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    found: set[str] = set()
    warnings: list[str] = []
    scanned_cells = 0
    m_cells = 0

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        scanned_cells += sum(1 for row in rows for value in row if value is not None)
        row_dates: dict[int, list[tuple[int, int, str]]] = {}
        col_dates: dict[int, list[tuple[int, int, str]]] = {}
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                parsed_date = _parse_date(value)
                if parsed_date is None:
                    continue
                row_dates.setdefault(row_index, []).append(
                    (row_index, col_index, parsed_date)
                )
                col_dates.setdefault(col_index, []).append(
                    (row_index, col_index, parsed_date)
                )

        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                if not _is_m_code(value):
                    continue
                m_cells += 1
                candidates = [
                    *row_dates.get(row_index, []),
                    *col_dates.get(col_index, []),
                ]
                if not candidates:
                    warnings.append(
                        f"{sheet.title}!{row_index + 1}:{col_index + 1}: cella M senza data riconoscibile."
                    )
                    continue
                found.add(_nearest_date(candidates, row_index, col_index))

    workbook.close()
    return WifeCalendarExcelImportResult(
        dates=sorted(found),
        scanned_cells=scanned_cells,
        m_cells=m_cells,
        warnings=warnings,
    )


def _is_m_code(value: Any) -> bool:
    return isinstance(value, str) and value.strip().upper() == "M"


def _nearest_date(
    candidates: list[tuple[int, int, str]], row_index: int, col_index: int
) -> str:
    def distance(candidate: tuple[int, int, str]) -> tuple[int, str]:
        candidate_row, candidate_col, parsed_date = candidate
        return (
            abs(candidate_row - row_index) + abs(candidate_col - col_index),
            parsed_date,
        )

    return min(candidates, key=distance)[2]


def _parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    if not normalized:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(normalized, fmt).date().isoformat()
        except ValueError:
            continue
    return None
